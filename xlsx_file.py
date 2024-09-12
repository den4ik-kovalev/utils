from collections import OrderedDict
from pathlib import Path
from typing import Union

import openpyxl


class XLSXFile:

    def __init__(self, path: Union[str, Path]):
        path = Path(path)
        if not path.suffix.startswith(".xls"):
            raise Exception("The file extension must be .xls")
        self._path = path

    @property
    def path(self) -> Path:
        return self._path

    def exists(self) -> bool:
        return self._path.exists()

    def read(self, sheet_idx: int = 0) -> list[OrderedDict]:

        wb = openpyxl.load_workbook(str(self._path))
        ws = wb.worksheets[sheet_idx]
        if ws.max_row == 0:
            return []

        keys = []
        for column in range(1, ws.max_column + 1):
            keys.append(ws.cell(1, column).value)

        data = []
        for row in range(2, ws.max_row + 1):
            dct = OrderedDict()
            for column, key in enumerate(keys, start=1):
                dct[key] = ws.cell(row, column).value
            data.append(dct)

        return data

    def write(self, data: list[OrderedDict], sheet_idx: int = 0) -> None:

        wb = openpyxl.Workbook()
        ws = wb.worksheets[sheet_idx]
        if not data:
            wb.save(str(self._path))
            return

        for column, key in enumerate(data[0].keys(), start=1):
            ws.cell(1, column, key)

        for row, dct in enumerate(data, start=2):
            for column, value in enumerate(dct.values(), start=1):
                ws.cell(row, column, value)

        self._adjust_columns_width(ws)
        wb.save(str(self._path))

    def delete(self) -> None:
        self._path.unlink()

    @staticmethod
    def _adjust_columns_width(sheet) -> None:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # adjusted_width = (max_length + 2) * 1.2
            adjusted_width = max_length
            sheet.column_dimensions[column].width = adjusted_width
